"""
Integrated Two-Stage Microgrid Optimization Framework
with Real-Time Sequential Dashboard Visualization.

Stage 1: Dynamic export/import limits calculation per time step.
Stage 2: ESS (battery) and HSH (hydrogen) scheduling using Stage 1 limits.

Each time step executes: Stage 1 → Stage 2 → Plot → Delay → Next time step.

Solver: PuLP (open-source, uses CBC by default).
        Gurobi can be used if a valid license is available by replacing
        `solver = pulp.PULP_CBC_CMD(msg=0)` with
        `solver = pulp.GUROBI(msg=0)`.
"""

import os
import sys
import time
import logging
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")          # non-interactive backend; switch to "TkAgg" for live GUI
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import FancyArrowPatch
import pulp
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")

# ============================================================
# SECTION 1 – CONFIGURATION
# ============================================================

# --- File paths ---
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "microgrid_30mins_data")
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
MG_FILES = {i: os.path.join(DATA_DIR, f"df_Mg{i}.csv") for i in range(1, 5)}
EXCEL_OUTPUT = os.path.join(OUTPUT_DIR, "optimization_results.xlsx")
LOG_FILE = os.path.join(OUTPUT_DIR, "optimization.log")
DASHBOARD_DIR = os.path.join(OUTPUT_DIR, "dashboard_frames")

# --- Capacity parameters (kW / kWh) ---
PV_CAPACITY = 365       # kW – rated PV capacity per MG reference
WT_CAPACITY = 347       # kW – rated WT capacity per MG reference
BESS_CAPACITY = 500     # kWh – battery energy storage
HYDRO_CAPACITY = 500    # kWh – hydrogen storage capacity

# --- Grid network constraints (kW) ---
GRID_EXPORT_MAX_TOTAL = 800   # max total export to grid across all MGs
GRID_IMPORT_MAX_TOTAL = 800   # max total import from grid across all MGs
MG_EXPORT_MAX = 300           # physical max export per MG
MG_IMPORT_MAX = 300           # physical max import per MG

# --- Stage 1 optimization parameters ---
ALPHA = 0.5     # export allocation factor
BETA = 0.7848   # import allocation factor
W1 = 0.2        # export deviation weight
W2 = 0.2        # import deviation weight

# --- Storage parameters ---
BATTERY_CHARGE_EFF = 0.95
BATTERY_DISCHARGE_EFF = 0.95
HYDROGEN_CHARGE_EFF = 0.95
HYDROGEN_DISCHARGE_EFF = 0.95

BATTERY_CHARGE_RATE = 300     # kW  max charge power
BATTERY_DISCHARGE_RATE = 300  # kW  max discharge power
HYDROGEN_CHARGE_RATE = 400    # kW  max electrolysis power
HYDROGEN_DISCHARGE_RATE = 200 # kW  max fuel-cell power

SOC_BESS_MIN = 0.1 * BESS_CAPACITY   # 10 % depth-of-discharge floor
SOC_BESS_MAX = 0.9 * BESS_CAPACITY   # 90 % ceiling
SOC_BESS_INIT = 0.5 * BESS_CAPACITY  # starting SOC

SOC_HYDRO_MIN = 0.05 * HYDRO_CAPACITY
SOC_HYDRO_MAX = 0.95 * HYDRO_CAPACITY
SOC_HYDRO_INIT = 0.5 * HYDRO_CAPACITY

# --- Optimization weights ---
W_SELF_SUFFICIENCY = 0.6    # Stage 2 – reward self-sufficiency
W_GRID_INTERACTION = 0.4    # Stage 2 – penalise grid interaction

# --- Prediction reliability factors (initial, updated each step) ---
PRED_RELIABILITY_INIT = 0.9  # 0–1
FORWARD_FACTOR_INIT = 1.0    # 0.5–2.0

# --- Visualization parameters ---
TIME_STEP_DELAY = 0.5         # seconds between time steps (reduced for non-interactive)
HISTORY_WINDOW = 5            # number of past steps shown in plots
SAVE_DASHBOARD_FRAMES = True  # save each dashboard frame as PNG

# ============================================================
# SECTION 2 – LOGGING SETUP
# ============================================================

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(DASHBOARD_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, mode="w"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger(__name__)


# ============================================================
# SECTION 3 – DATA LOADING
# ============================================================

def load_microgrid_data() -> dict[int, pd.DataFrame]:
    """Load CSV data for all four microgrids.

    Returns
    -------
    dict[int, pd.DataFrame]
        Keys 1–4 mapping to per-MG DataFrames.
    """
    mg_data: dict[int, pd.DataFrame] = {}
    required_cols = [
        "PV_Predicted", "PV_Actual",
        "WT_Predicted", "WT_Actual",
        "ED_Predicted", "ED_Actual",
        "HD_Predicted", "HD_Actual",
        "TAS94", "FIT",
    ]
    for mg_id, filepath in MG_FILES.items():
        if not os.path.exists(filepath):
            raise FileNotFoundError(
                f"Data file not found: {filepath}\n"
                "Please ensure microgrid_30mins_data/ contains df_Mg1.csv … df_Mg4.csv"
            )
        df = pd.read_csv(filepath, parse_dates=["DateTime"])
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            raise ValueError(f"MG{mg_id} is missing columns: {missing}")
        df = df.sort_values("DateTime").reset_index(drop=True)
        mg_data[mg_id] = df
        logger.info("Loaded MG%d: %d time steps (%s → %s)",
                    mg_id, len(df),
                    df["DateTime"].iloc[0], df["DateTime"].iloc[-1])
    return mg_data


# ============================================================
# SECTION 4 – STAGE 1 OPTIMISATION: DYNAMIC LIMITS
# ============================================================

def optimize_stage_1_dynamic_limits(
    t: int,
    mg_data: dict[int, pd.DataFrame],
    pred_reliability: dict[int, float],
    forward_factor: dict[int, float],
) -> dict:
    """Calculate dynamic export/import limits for each MG at time step *t*.

    The optimization minimizes weighted deviations from grid export/import
    targets while respecting:
      - physical per-MG limits
      - grid network totals
      - prediction-reliability-scaled dynamic adjustments

    Parameters
    ----------
    t              : current time-step index
    mg_data        : loaded CSV data (keys 1–4)
    pred_reliability : reliability factor per MG (0–1)
    forward_factor   : forward-looking generation/demand factor (0.5–2.0)

    Returns
    -------
    dict with keys
        export_max, export_min, import_max, import_min  – per MG (keys 1–4)
        solve_status, objective_value, grid_utilisation
    """
    n_mg = len(mg_data)      # 4 microgrids

    # ---- build per-MG net generation at this time step ----
    net_gen: dict[int, float] = {}
    for mg_id, df in mg_data.items():
        row = df.iloc[t]
        pv_p = float(row["PV_Predicted"])
        wt_p = float(row["WT_Predicted"])
        ed_p = float(row["ED_Predicted"])
        hd_p = float(row["HD_Predicted"])
        rf = pred_reliability[mg_id]
        ff = forward_factor[mg_id]
        net_gen[mg_id] = rf * ff * (pv_p + wt_p) - (ed_p + hd_p)

    # ---- PuLP model ----
    model = pulp.LpProblem("Stage1_DynamicLimits", pulp.LpMinimize)

    # Decision variables: export / import limit adjustments per MG
    ex_max = {i: pulp.LpVariable(f"ex_max_{i}", 0, MG_EXPORT_MAX) for i in mg_data}
    ex_min = {i: pulp.LpVariable(f"ex_min_{i}", 0, MG_EXPORT_MAX) for i in mg_data}
    im_max = {i: pulp.LpVariable(f"im_max_{i}", 0, MG_IMPORT_MAX) for i in mg_data}
    im_min = {i: pulp.LpVariable(f"im_min_{i}", 0, MG_IMPORT_MAX) for i in mg_data}

    # Slack variables for soft-constraint deviations
    dev_ex = {i: pulp.LpVariable(f"dev_ex_{i}", 0) for i in mg_data}
    dev_im = {i: pulp.LpVariable(f"dev_im_{i}", 0) for i in mg_data}

    # ---- Objective: minimise total deviation from targets ----
    target_ex = {
        i: max(0.0, ALPHA * net_gen[i]) for i in mg_data
    }
    target_im = {
        i: max(0.0, -BETA * net_gen[i]) for i in mg_data
    }
    model += (
        pulp.lpSum(W1 * dev_ex[i] + W2 * dev_im[i] for i in mg_data)
    ), "obj"

    for i in mg_data:
        rf = pred_reliability[i]
        ff = forward_factor[i]

        # Dynamic max export: reliability-scaled physical limit
        dyn_ex_max = rf * ff * MG_EXPORT_MAX
        dyn_im_max = (1 - rf * (ff - 1) / 2) * MG_IMPORT_MAX  # slight bias

        # Export limit bounds
        model += ex_max[i] <= dyn_ex_max,    f"ex_max_upper_{i}"
        model += ex_min[i] <= ex_max[i],     f"ex_min_le_max_{i}"

        # Import limit bounds
        model += im_max[i] <= dyn_im_max,    f"im_max_upper_{i}"
        model += im_min[i] <= im_max[i],     f"im_min_le_max_{i}"

        # Soft deviation from targets
        model += ex_max[i] >= target_ex[i] - dev_ex[i], f"ex_target_{i}"
        model += im_max[i] >= target_im[i] - dev_im[i], f"im_target_{i}"

    # Grid network constraint: sum of max exports/imports ≤ grid capacity
    model += pulp.lpSum(ex_max[i] for i in mg_data) <= GRID_EXPORT_MAX_TOTAL, "grid_ex"
    model += pulp.lpSum(im_max[i] for i in mg_data) <= GRID_IMPORT_MAX_TOTAL, "grid_im"

    # ---- Solve ----
    solver = pulp.PULP_CBC_CMD(msg=0)
    status = model.solve(solver)
    status_str = pulp.LpStatus[model.status]

    if model.status != pulp.constants.LpStatusOptimal:
        logger.warning("Stage 1 t=%d: solver returned '%s' – using fallback limits.", t, status_str)
        # Fallback: use proportional allocation
        return _stage1_fallback(t, mg_data, pred_reliability, forward_factor, net_gen)

    result = {
        "export_max":  {i: float(ex_max[i].value() or 0) for i in mg_data},
        "export_min":  {i: float(ex_min[i].value() or 0) for i in mg_data},
        "import_max":  {i: float(im_max[i].value() or 0) for i in mg_data},
        "import_min":  {i: float(im_min[i].value() or 0) for i in mg_data},
        "net_gen":     net_gen,
        "solve_status": status_str,
        "objective_value": pulp.value(model.objective),
        "grid_export_utilisation": (
            sum(float(ex_max[i].value() or 0) for i in mg_data) / GRID_EXPORT_MAX_TOTAL * 100
        ),
        "grid_import_utilisation": (
            sum(float(im_max[i].value() or 0) for i in mg_data) / GRID_IMPORT_MAX_TOTAL * 100
        ),
    }
    return result


def _stage1_fallback(t, mg_data, pred_reliability, forward_factor, net_gen):
    """Simple proportional allocation used when Stage 1 solver is infeasible."""
    export_max, export_min, import_max, import_min = {}, {}, {}, {}
    total_pos = max(1e-6, sum(max(0, v) for v in net_gen.values()))
    total_neg = max(1e-6, sum(max(0, -v) for v in net_gen.values()))

    for i in mg_data:
        rf = pred_reliability[i]
        share_ex = max(0, net_gen[i]) / total_pos
        share_im = max(0, -net_gen[i]) / total_neg
        export_max[i] = min(rf * MG_EXPORT_MAX, ALPHA * GRID_EXPORT_MAX_TOTAL * share_ex)
        export_min[i] = export_max[i] * 0.1
        import_max[i] = min(rf * MG_IMPORT_MAX, BETA * GRID_IMPORT_MAX_TOTAL * share_im)
        import_min[i] = import_max[i] * 0.1

    return {
        "export_max": export_max,
        "export_min": export_min,
        "import_max": import_max,
        "import_min": import_min,
        "net_gen": net_gen,
        "solve_status": "Fallback",
        "objective_value": None,
        "grid_export_utilisation": sum(export_max.values()) / GRID_EXPORT_MAX_TOTAL * 100,
        "grid_import_utilisation": sum(import_max.values()) / GRID_IMPORT_MAX_TOTAL * 100,
    }


# ============================================================
# SECTION 5 – STAGE 2 OPTIMISATION: ESS / HSH SCHEDULING
# ============================================================

def optimize_stage_2_ess_hsh_scheduling(
    t: int,
    mg_data: dict[int, pd.DataFrame],
    stage1_result: dict,
    bess_soc: float,
    hydro_soc: float,
) -> dict:
    """Optimise battery + hydrogen scheduling for the *aggregate* of all MGs.

    The aggregated system has:
      - Total PV/WT generation (predicted)
      - Total electricity/hydrogen demand (predicted)
      - One BESS bank and one HSH tank shared among the MGs
      - Grid import/export bounded by Stage 1 limits (sum over MGs)

    Parameters
    ----------
    t              : current time-step index
    mg_data        : loaded CSV data
    stage1_result  : output of optimize_stage_1_dynamic_limits()
    bess_soc       : battery SOC at start of this step (kWh)
    hydro_soc      : hydrogen SOC at start of this step (kWh)

    Returns
    -------
    dict with scheduling results and updated SOCs.
    """
    # ---- aggregate predicted generation / demand ----
    pv_pred = sum(float(df.iloc[t]["PV_Predicted"]) for df in mg_data.values())
    wt_pred = sum(float(df.iloc[t]["WT_Predicted"]) for df in mg_data.values())
    ed_pred = sum(float(df.iloc[t]["ED_Predicted"]) for df in mg_data.values())
    hd_pred = sum(float(df.iloc[t]["HD_Predicted"]) for df in mg_data.values())

    pv_act  = sum(float(df.iloc[t]["PV_Actual"])    for df in mg_data.values())
    wt_act  = sum(float(df.iloc[t]["WT_Actual"])    for df in mg_data.values())
    ed_act  = sum(float(df.iloc[t]["ED_Actual"])    for df in mg_data.values())
    hd_act  = sum(float(df.iloc[t]["HD_Actual"])    for df in mg_data.values())

    tariff  = float(list(mg_data.values())[0].iloc[t]["TAS94"])
    fit     = float(list(mg_data.values())[0].iloc[t]["FIT"])

    # ---- limits from Stage 1 (aggregate) ----
    ex_max_agg = sum(stage1_result["export_max"].values())
    ex_min_agg = sum(stage1_result["export_min"].values())
    im_max_agg = sum(stage1_result["import_max"].values())
    im_min_agg = sum(stage1_result["import_min"].values())

    # ---- PuLP model ----
    model = pulp.LpProblem("Stage2_ESSScheduling", pulp.LpMinimize)

    # Decision variables
    p_bess_ch  = pulp.LpVariable("p_bess_ch",  0, BATTERY_CHARGE_RATE)
    p_bess_dis = pulp.LpVariable("p_bess_dis", 0, BATTERY_DISCHARGE_RATE)
    p_hyd_ch   = pulp.LpVariable("p_hyd_ch",  0, HYDROGEN_CHARGE_RATE)
    p_hyd_dis  = pulp.LpVariable("p_hyd_dis", 0, HYDROGEN_DISCHARGE_RATE)
    p_grid_ex  = pulp.LpVariable("p_grid_ex", 0, ex_max_agg)
    p_grid_im  = pulp.LpVariable("p_grid_im", 0, im_max_agg)

    # Soft-constraint slack variables for energy balance (avoids infeasibility)
    slack_plus  = pulp.LpVariable("slack_plus",  0)
    slack_minus = pulp.LpVariable("slack_minus", 0)

    bess_soc_new  = pulp.LpVariable("bess_soc_new",  SOC_BESS_MIN,  SOC_BESS_MAX)
    hydro_soc_new = pulp.LpVariable("hydro_soc_new", SOC_HYDRO_MIN, SOC_HYDRO_MAX)

    BIG_M_SLACK = 1e4   # large penalty for violating energy balance

    # ---- Objective ----
    # Maximise self-sufficiency (minimise grid use + import cost - export revenue)
    # Penalise slack heavily to enforce energy balance
    model += (
        W_GRID_INTERACTION * (tariff * p_grid_im - fit * p_grid_ex)
        + W_SELF_SUFFICIENCY * (p_grid_im + p_grid_ex)
        + BIG_M_SLACK * (slack_plus + slack_minus)
    ), "obj"

    # ---- Energy balance (electricity): generation + import + discharge = demand + charge + export
    generation = pv_pred + wt_pred
    model += (
        generation
        + p_grid_im
        + BATTERY_DISCHARGE_EFF * p_bess_dis
        + HYDROGEN_DISCHARGE_EFF * p_hyd_dis
        + slack_plus
        == ed_pred
        + p_bess_ch
        + p_hyd_ch
        + p_grid_ex
        + slack_minus
    ), "energy_balance"

    # ---- Hydrogen demand: HSH must supply hydrogen demand
    # (converted from kWh equivalent with factor 1 for simplicity)
    model += p_hyd_dis >= hd_pred, "hyd_demand"

    # ---- BESS SOC dynamics ----
    dt = 0.5   # 30-minute time step
    model += (
        bess_soc_new == bess_soc
        + BATTERY_CHARGE_EFF * p_bess_ch * dt
        - (1 / BATTERY_DISCHARGE_EFF) * p_bess_dis * dt
    ), "bess_soc_update"

    # ---- Hydrogen SOC dynamics ----
    model += (
        hydro_soc_new == hydro_soc
        + HYDROGEN_CHARGE_EFF * p_hyd_ch * dt
        - (1 / HYDROGEN_DISCHARGE_EFF) * p_hyd_dis * dt
    ), "hydro_soc_update"

    # ---- Minimum grid limits ----
    model += p_grid_ex >= ex_min_agg, "ex_min"
    model += p_grid_im >= im_min_agg, "im_min"

    # ---- Solve ----
    solver = pulp.PULP_CBC_CMD(msg=0)
    status = model.solve(solver)
    status_str = pulp.LpStatus[model.status]

    if model.status != pulp.constants.LpStatusOptimal:
        logger.warning("Stage 2 t=%d: solver '%s' – using heuristic scheduling.", t, status_str)
        return _stage2_fallback(
            t, pv_pred, wt_pred, ed_pred, hd_pred,
            pv_act, wt_act, ed_act, hd_act,
            bess_soc, hydro_soc, ex_max_agg, im_max_agg, tariff, fit, status_str
        )

    def _v(var):
        return float(var.value() or 0)

    bess_soc_out  = _v(bess_soc_new)
    hydro_soc_out = _v(hydro_soc_new)
    p_bess_ch_v   = _v(p_bess_ch)
    p_bess_dis_v  = _v(p_bess_dis)
    p_hyd_ch_v    = _v(p_hyd_ch)
    p_hyd_dis_v   = _v(p_hyd_dis)
    p_grid_ex_v   = _v(p_grid_ex)
    p_grid_im_v   = _v(p_grid_im)
    slack_plus_v  = _v(slack_plus)
    slack_minus_v = _v(slack_minus)

    energy_balance = (
        pv_pred + wt_pred
        + p_grid_im_v
        + BATTERY_DISCHARGE_EFF * p_bess_dis_v
        + HYDROGEN_DISCHARGE_EFF * p_hyd_dis_v
        - ed_pred - p_bess_ch_v - p_hyd_ch_v - p_grid_ex_v
    )

    return {
        "pv_predicted": pv_pred,
        "pv_actual":    pv_act,
        "wt_predicted": wt_pred,
        "wt_actual":    wt_act,
        "ed_predicted": ed_pred,
        "ed_actual":    ed_act,
        "hd_predicted": hd_pred,
        "hd_actual":    hd_act,
        "tariff":       tariff,
        "fit":          fit,
        # BESS
        "bess_charge":    p_bess_ch_v,
        "bess_discharge": p_bess_dis_v,
        "bess_soc":       bess_soc_out,
        # Hydrogen
        "hyd_charge":     p_hyd_ch_v,
        "hyd_discharge":  p_hyd_dis_v,
        "hyd_soc":        hydro_soc_out,
        # Grid
        "grid_export":    p_grid_ex_v,
        "grid_import":    p_grid_im_v,
        # Balance
        "energy_balance": energy_balance,
        "slack_plus":     slack_plus_v,
        "slack_minus":    slack_minus_v,
        "solve_status":   status_str,
        "objective_value": pulp.value(model.objective),
    }


def _stage2_fallback(
    t, pv_pred, wt_pred, ed_pred, hd_pred,
    pv_act, wt_act, ed_act, hd_act,
    bess_soc, hydro_soc, ex_max_agg, im_max_agg, tariff, fit, status_str
):
    """Rule-based heuristic scheduling when Stage 2 solver is infeasible."""
    dt = 0.5
    net = pv_pred + wt_pred - ed_pred - hd_pred

    p_hyd_dis = min(hd_pred, HYDROGEN_DISCHARGE_RATE,
                    (hydro_soc - SOC_HYDRO_MIN) / (dt / HYDROGEN_DISCHARGE_EFF))
    p_hyd_ch  = 0.0

    hydro_soc_new = hydro_soc + HYDROGEN_CHARGE_EFF * p_hyd_ch * dt - (1 / HYDROGEN_DISCHARGE_EFF) * p_hyd_dis * dt
    hydro_soc_new = np.clip(hydro_soc_new, SOC_HYDRO_MIN, SOC_HYDRO_MAX)

    residual = net + p_hyd_dis * HYDROGEN_DISCHARGE_EFF
    if residual > 0:
        p_bess_ch  = min(residual, BATTERY_CHARGE_RATE, (SOC_BESS_MAX - bess_soc) / (BATTERY_CHARGE_EFF * dt))
        p_bess_dis = 0.0
    else:
        p_bess_ch  = 0.0
        p_bess_dis = min(-residual, BATTERY_DISCHARGE_RATE, (bess_soc - SOC_BESS_MIN) * BATTERY_DISCHARGE_EFF / dt)

    bess_soc_new = bess_soc + BATTERY_CHARGE_EFF * p_bess_ch * dt - (1 / BATTERY_DISCHARGE_EFF) * p_bess_dis * dt
    bess_soc_new = np.clip(bess_soc_new, SOC_BESS_MIN, SOC_BESS_MAX)

    residual2 = residual - p_bess_ch + BATTERY_DISCHARGE_EFF * p_bess_dis
    p_grid_ex = max(0, min(residual2, ex_max_agg))
    p_grid_im = max(0, min(-residual2, im_max_agg))

    energy_balance = (
        pv_pred + wt_pred + p_grid_im
        + BATTERY_DISCHARGE_EFF * p_bess_dis
        + HYDROGEN_DISCHARGE_EFF * p_hyd_dis
        - ed_pred - p_bess_ch - p_hyd_ch - p_grid_ex
    )

    return {
        "pv_predicted": pv_pred, "pv_actual": pv_act,
        "wt_predicted": wt_pred, "wt_actual": wt_act,
        "ed_predicted": ed_pred, "ed_actual": ed_act,
        "hd_predicted": hd_pred, "hd_actual": hd_act,
        "tariff": tariff, "fit": fit,
        "bess_charge": p_bess_ch, "bess_discharge": p_bess_dis, "bess_soc": bess_soc_new,
        "hyd_charge": p_hyd_ch,   "hyd_discharge": p_hyd_dis,   "hyd_soc": hydro_soc_new,
        "grid_export": p_grid_ex, "grid_import": p_grid_im,
        "energy_balance": energy_balance, "slack_plus": 0, "slack_minus": 0,
        "solve_status": status_str, "objective_value": None,
    }


# ============================================================
# SECTION 6 – REAL-TIME DASHBOARD
# ============================================================

def create_real_time_dashboard(
    t: int,
    mg_data: dict[int, pd.DataFrame],
    stage1_result: dict,
    stage2_result: dict,
    history: list[dict],
    fig: plt.Figure | None,
    axes_dict: dict | None,
) -> tuple[plt.Figure, dict]:
    """Build or update the real-time two-section dashboard.

    Parameters
    ----------
    t            : current time-step index
    mg_data      : loaded CSV data
    stage1_result: Stage 1 outputs
    stage2_result: Stage 2 outputs
    history      : list of previous combined results
    fig          : existing Figure (None to create new)
    axes_dict    : dict of named Axes (None to create new)

    Returns
    -------
    (fig, axes_dict)
    """
    n_mg = len(mg_data)

    # Colour scheme (defined first so constants are available throughout)
    DARK_BG    = "#0d1117"
    PANEL_BG   = "#161b22"
    TEXT_COL   = "#e6edf3"
    ACCENT     = "#58a6ff"
    GREEN      = "#3fb950"
    RED        = "#f85149"
    ORANGE     = "#d29922"
    PURPLE     = "#bc8cff"
    CYAN       = "#39d353"
    BORDER_COL = "#30363d"

    if fig is None:
        fig = plt.figure(figsize=(22, 14), facecolor=DARK_BG)
        fig.suptitle(
            "Microgrid Two-Stage Optimization Dashboard",
            fontsize=16, color="white", fontweight="bold", y=0.98,
        )
    else:
        fig.clear()
        fig.set_facecolor(DARK_BG)
        fig.suptitle(
            "Microgrid Two-Stage Optimization Dashboard",
            fontsize=16, color="white", fontweight="bold", y=0.98,
        )

    # ---- Layout: 4 rows × 6 cols ----
    gs = gridspec.GridSpec(
        4, 6, figure=fig,
        left=0.06, right=0.97,
        top=0.93, bottom=0.06,
        hspace=0.55, wspace=0.45,
    )

    def style_ax(ax, title=""):
        ax.set_facecolor(PANEL_BG)
        for spine in ax.spines.values():
            spine.set_edgecolor(BORDER_COL)
        ax.tick_params(colors=TEXT_COL, labelsize=7)
        ax.xaxis.label.set_color(TEXT_COL)
        ax.yaxis.label.set_color(TEXT_COL)
        if title:
            ax.set_title(title, color=ACCENT, fontsize=8, fontweight="bold", pad=4)

    # -----------------------------------------------------------------
    # Section 1 header banner
    # -----------------------------------------------------------------
    ax_hdr1 = fig.add_subplot(gs[0, :])
    ax_hdr1.set_facecolor(PANEL_BG)
    ts_label = (
        mg_data[1].iloc[t]["DateTime"].strftime("%H:%M")
        if hasattr(mg_data[1].iloc[t]["DateTime"], "strftime")
        else f"t={t}"
    )
    ax_hdr1.text(
        0.02, 0.5,
        f"▶  STAGE 1 – Dynamic Limits    [Time Step {t:02d}  {ts_label}]",
        transform=ax_hdr1.transAxes,
        fontsize=11, color=ACCENT, fontweight="bold", va="center",
    )
    ax_hdr1.text(
        0.55, 0.5,
        (f"Grid Export Util: {stage1_result['grid_export_utilisation']:.1f}%   "
         f"Grid Import Util: {stage1_result['grid_import_utilisation']:.1f}%   "
         f"Solver: {stage1_result['solve_status']}"),
        transform=ax_hdr1.transAxes,
        fontsize=9, color=TEXT_COL, va="center",
    )
    ax_hdr1.axis("off")
    for spine in ax_hdr1.spines.values():
        spine.set_edgecolor(BORDER_COL)

    # -----------------------------------------------------------------
    # Stage 1 – Per-MG export/import bars  (row 1, cols 0-2)
    # -----------------------------------------------------------------
    ax_ex = fig.add_subplot(gs[1, :3])
    style_ax(ax_ex, "Export Limits per MG (kW)")

    mg_ids = list(stage1_result["export_max"].keys())
    x = np.arange(len(mg_ids))
    w = 0.35
    ex_max_vals = [stage1_result["export_max"][i] for i in mg_ids]
    ex_min_vals = [stage1_result["export_min"][i] for i in mg_ids]
    bars_ex_max = ax_ex.bar(x - w/2, ex_max_vals, w, color=GREEN,  alpha=0.8, label="Export Max")
    bars_ex_min = ax_ex.bar(x + w/2, ex_min_vals, w, color=CYAN,   alpha=0.8, label="Export Min")
    # Physical limit reference line
    ax_ex.axhline(MG_EXPORT_MAX, color=RED, linestyle="--", linewidth=1, label="Physical Max")
    ax_ex.set_xticks(x)
    ax_ex.set_xticklabels([f"MG{i}" for i in mg_ids], color=TEXT_COL, fontsize=8)
    ax_ex.legend(fontsize=7, facecolor=PANEL_BG, labelcolor=TEXT_COL,
                 loc="upper right", framealpha=0.6)
    ax_ex.set_ylabel("Power (kW)", fontsize=7)

    ax_im = fig.add_subplot(gs[1, 3:])
    style_ax(ax_im, "Import Limits per MG (kW)")

    im_max_vals = [stage1_result["import_max"][i] for i in mg_ids]
    im_min_vals = [stage1_result["import_min"][i] for i in mg_ids]
    ax_im.bar(x - w/2, im_max_vals, w, color=ORANGE, alpha=0.8, label="Import Max")
    ax_im.bar(x + w/2, im_min_vals, w, color=PURPLE, alpha=0.8, label="Import Min")
    ax_im.axhline(MG_IMPORT_MAX, color=RED, linestyle="--", linewidth=1, label="Physical Max")
    ax_im.set_xticks(x)
    ax_im.set_xticklabels([f"MG{i}" for i in mg_ids], color=TEXT_COL, fontsize=8)
    ax_im.legend(fontsize=7, facecolor=PANEL_BG, labelcolor=TEXT_COL,
                 loc="upper right", framealpha=0.6)
    ax_im.set_ylabel("Power (kW)", fontsize=7)

    # Net energy indicator text
    net_labels = []
    for i in mg_ids:
        ng = stage1_result["net_gen"][i]
        label = f"MG{i}: {'↑Exporter' if ng >= 0 else '↓Importer'} ({ng:+.0f} kW)"
        net_labels.append(label)
    ax_im.text(
        0.01, -0.28, "  |  ".join(net_labels),
        transform=ax_im.transAxes, fontsize=7, color=TEXT_COL,
    )

    # -----------------------------------------------------------------
    # Section 2 header banner
    # -----------------------------------------------------------------
    ax_hdr2 = fig.add_subplot(gs[2, :])
    ax_hdr2.set_facecolor(PANEL_BG)
    ax_hdr2.text(
        0.02, 0.5,
        "▶  STAGE 2 – ESS / HSH Scheduling",
        transform=ax_hdr2.transAxes,
        fontsize=11, color=GREEN, fontweight="bold", va="center",
    )
    s2 = stage2_result
    balance_str = f"Energy Balance: {s2['energy_balance']:+.2f} kW"
    balance_col = GREEN if abs(s2["energy_balance"]) < 1 else RED
    ax_hdr2.text(
        0.55, 0.5,
        (f"{balance_str}   |   Grid Import: {s2['grid_import']:.1f} kW   "
         f"Grid Export: {s2['grid_export']:.1f} kW   "
         f"Solver: {s2['solve_status']}"),
        transform=ax_hdr2.transAxes,
        fontsize=9, color=balance_col, va="center",
    )
    ax_hdr2.axis("off")
    for spine in ax_hdr2.spines.values():
        spine.set_edgecolor(BORDER_COL)

    # -----------------------------------------------------------------
    # Stage 2 – Generation panel  (row 3, cols 0-1)
    # -----------------------------------------------------------------
    ax_gen = fig.add_subplot(gs[3, :2])
    style_ax(ax_gen, "Generation & Demand (Predicted vs Actual, kW)")

    categories = ["PV", "WT", "ED", "HD"]
    pred_vals   = [s2["pv_predicted"], s2["wt_predicted"], s2["ed_predicted"], s2["hd_predicted"]]
    actual_vals = [s2["pv_actual"],    s2["wt_actual"],    s2["ed_actual"],    s2["hd_actual"]]
    x2 = np.arange(len(categories))
    ax_gen.bar(x2 - 0.2, pred_vals,   0.35, color=ACCENT,  alpha=0.8, label="Predicted")
    ax_gen.bar(x2 + 0.2, actual_vals, 0.35, color=ORANGE,  alpha=0.8, label="Actual")
    ax_gen.set_xticks(x2)
    ax_gen.set_xticklabels(categories, color=TEXT_COL, fontsize=8)
    ax_gen.legend(fontsize=7, facecolor=PANEL_BG, labelcolor=TEXT_COL, framealpha=0.6)
    ax_gen.set_ylabel("kW", fontsize=7)

    # -----------------------------------------------------------------
    # Stage 2 – BESS panel  (row 3, cols 2-3)
    # -----------------------------------------------------------------
    ax_bess = fig.add_subplot(gs[3, 2:4])
    style_ax(ax_bess, "Battery Storage (kWh / kW)")

    # Historical SOC trace (use step indices from history for consistency)
    recent_hist    = history[-HISTORY_WINDOW:]
    hist_t         = [h["step"] for h in recent_hist]
    hist_bess_soc  = [h["stage2"]["bess_soc"]       for h in recent_hist]
    hist_bess_ch   = [h["stage2"]["bess_charge"]     for h in recent_hist]
    hist_bess_dis  = [h["stage2"]["bess_discharge"]  for h in recent_hist]

    ax_bess2 = ax_bess.twinx()
    ax_bess2.set_facecolor(PANEL_BG)
    ax_bess2.tick_params(colors=TEXT_COL, labelsize=7)

    if hist_t:
        ax_bess.plot(hist_t, hist_bess_soc, color=ACCENT, linewidth=1.5, label="SOC (kWh)")
        ax_bess2.bar(
            [ti - 0.2 for ti in hist_t], hist_bess_ch,  0.35,
            color=GREEN, alpha=0.7, label="Charge"
        )
        ax_bess2.bar(
            [ti + 0.2 for ti in hist_t], hist_bess_dis, 0.35,
            color=RED,   alpha=0.7, label="Discharge"
        )

    # Current step marker
    ax_bess.scatter([t], [s2["bess_soc"]], color="white", zorder=5, s=40)
    ax_bess.axhline(SOC_BESS_MAX, color=GREEN, linestyle="--", linewidth=0.8, alpha=0.5)
    ax_bess.axhline(SOC_BESS_MIN, color=RED,   linestyle="--", linewidth=0.8, alpha=0.5)
    ax_bess.set_ylabel("SOC (kWh)", fontsize=7, color=ACCENT)
    ax_bess2.set_ylabel("Power (kW)", fontsize=7, color=GREEN)
    ax_bess.legend(fontsize=7, facecolor=PANEL_BG, labelcolor=TEXT_COL,
                   loc="upper left", framealpha=0.6)
    ax_bess2.legend(fontsize=7, facecolor=PANEL_BG, labelcolor=TEXT_COL,
                    loc="upper right", framealpha=0.6)

    # -----------------------------------------------------------------
    # Stage 2 – Hydrogen panel  (row 3, cols 4-5)
    # -----------------------------------------------------------------
    ax_hyd = fig.add_subplot(gs[3, 4:])
    style_ax(ax_hyd, "Hydrogen Storage (kWh / kW)")

    hist_hyd_soc = [h["stage2"]["hyd_soc"]       for h in recent_hist]
    hist_hyd_ch  = [h["stage2"]["hyd_charge"]     for h in recent_hist]
    hist_hyd_dis = [h["stage2"]["hyd_discharge"]  for h in recent_hist]

    ax_hyd2 = ax_hyd.twinx()
    ax_hyd2.set_facecolor(PANEL_BG)
    ax_hyd2.tick_params(colors=TEXT_COL, labelsize=7)

    if hist_t:
        ax_hyd.plot(hist_t, hist_hyd_soc, color=PURPLE,  linewidth=1.5, label="SOC (kWh)")
        ax_hyd2.bar(
            [ti - 0.2 for ti in hist_t], hist_hyd_ch,  0.35,
            color=CYAN,  alpha=0.7, label="Charge"
        )
        ax_hyd2.bar(
            [ti + 0.2 for ti in hist_t], hist_hyd_dis, 0.35,
            color=ORANGE, alpha=0.7, label="Discharge"
        )

    ax_hyd.scatter([t], [s2["hyd_soc"]], color="white", zorder=5, s=40)
    ax_hyd.axhline(SOC_HYDRO_MAX, color=CYAN,  linestyle="--", linewidth=0.8, alpha=0.5)
    ax_hyd.axhline(SOC_HYDRO_MIN, color=RED,   linestyle="--", linewidth=0.8, alpha=0.5)
    ax_hyd.set_ylabel("SOC (kWh)", fontsize=7, color=PURPLE)
    ax_hyd2.set_ylabel("Power (kW)", fontsize=7, color=CYAN)
    ax_hyd.legend(fontsize=7, facecolor=PANEL_BG, labelcolor=TEXT_COL,
                  loc="upper left", framealpha=0.6)
    ax_hyd2.legend(fontsize=7, facecolor=PANEL_BG, labelcolor=TEXT_COL,
                   loc="upper right", framealpha=0.6)

    plt.pause(0.001)

    if SAVE_DASHBOARD_FRAMES:
        frame_path = os.path.join(DASHBOARD_DIR, f"step_{t:03d}.png")
        fig.savefig(frame_path, dpi=120, facecolor=DARK_BG, bbox_inches="tight")
        logger.debug("Dashboard frame saved: %s", frame_path)

    return fig, {}


# ============================================================
# SECTION 7 – RESULTS EXPORT
# ============================================================

def save_results_to_excel(history: list[dict], output_path: str) -> None:
    """Save all optimization results to an Excel workbook with multiple sheets.

    Parameters
    ----------
    history     : list of per-time-step combined result dicts
    output_path : destination .xlsx path
    """
    if not history:
        logger.warning("No results to save.")
        return

    n_mg  = 4
    n_t   = len(history)
    times = [h["time"] for h in history]

    def _df_s1(key, sub_key):
        """Build DataFrame from Stage 1 per-MG values."""
        data = {f"MG{i}": [h["stage1"][key][i] for h in history] for i in range(1, n_mg + 1)}
        return pd.DataFrame(data, index=times)

    def _series_s2(key):
        return pd.Series([h["stage2"][key] for h in history], index=times, name=key)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    def _write_df(sheet_name, df):
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["DateTime"] + list(df.columns))
        for row in df.itertuples():
            ws.append(list(row))

    # Stage 1 sheets
    _write_df("Stage1_Export_Max", _df_s1("export_max", None))
    _write_df("Stage1_Export_Min", _df_s1("export_min", None))
    _write_df("Stage1_Import_Max", _df_s1("import_max", None))
    _write_df("Stage1_Import_Min", _df_s1("import_min", None))

    # Stage 2 sheets – scalar series → single-column DataFrames
    s2_keys = [
        ("Stage2_Battery_SOC",      "bess_soc"),
        ("Stage2_Battery_Power_Ch", "bess_charge"),
        ("Stage2_Battery_Power_Di", "bess_discharge"),
        ("Stage2_Hydrogen_SOC",     "hyd_soc"),
        ("Stage2_Hydrogen_Pwr_Ch",  "hyd_charge"),
        ("Stage2_Hydrogen_Pwr_Di",  "hyd_discharge"),
        ("Stage2_Grid_Import",      "grid_import"),
        ("Stage2_Grid_Export",      "grid_export"),
        ("Stage2_Energy_Balance",   "energy_balance"),
    ]
    for sheet_name, key in s2_keys:
        s = _series_s2(key)
        df = s.to_frame()
        _write_df(sheet_name, df)

    # Summary statistics sheet
    summary_rows = [
        ["Metric", "Min", "Max", "Mean", "Sum"],
    ]
    for _, key in s2_keys:
        vals = [h["stage2"][key] for h in history]
        summary_rows.append([
            key,
            round(min(vals), 4),
            round(max(vals), 4),
            round(np.mean(vals), 4),
            round(sum(vals), 4),
        ])

    ws_sum = wb.create_sheet(title="Summary")
    for row in summary_rows:
        ws_sum.append(row)

    wb.save(output_path)
    logger.info("Results saved to %s", output_path)


# ============================================================
# SECTION 8 – MAIN EXECUTION LOOP
# ============================================================

def run_optimization(max_steps: int | None = None, interactive_plot: bool = False) -> list[dict]:
    """Execute the two-stage optimization for every time step in the data.

    Parameters
    ----------
    max_steps        : limit number of time steps (None = all)
    interactive_plot : if True, render live matplotlib window (requires GUI)

    Returns
    -------
    history : list of per-step result dicts
    """
    logger.info("=" * 70)
    logger.info("Integrated Microgrid Optimization  –  start %s", datetime.now())
    logger.info("=" * 70)

    # ---- Load data ----
    mg_data = load_microgrid_data()
    n_steps = len(next(iter(mg_data.values())))
    if max_steps is not None:
        n_steps = min(n_steps, max_steps)
    logger.info("Running %d time steps.", n_steps)

    # ---- Initialise state ----
    bess_soc  = SOC_BESS_INIT
    hydro_soc = SOC_HYDRO_INIT

    pred_reliability = {i: PRED_RELIABILITY_INIT  for i in mg_data}
    forward_factor   = {i: FORWARD_FACTOR_INIT    for i in mg_data}

    history: list[dict] = []
    fig  = None
    axes = None

    if interactive_plot:
        plt.ion()

    for t in range(n_steps):
        t_start = time.time()
        timestamp = mg_data[1].iloc[t]["DateTime"]
        logger.info("──── Time Step %d / %d  (%s) ────", t, n_steps - 1, timestamp)

        # ---- Stage 1 ----
        s1 = optimize_stage_1_dynamic_limits(t, mg_data, pred_reliability, forward_factor)
        logger.info(
            "Stage 1  status=%-10s  obj=%-10s  ExUtil=%.1f%%  ImUtil=%.1f%%",
            s1["solve_status"],
            f"{s1['objective_value']:.4f}" if s1["objective_value"] is not None else "N/A",
            s1["grid_export_utilisation"],
            s1["grid_import_utilisation"],
        )

        # ---- Stage 2 ----
        s2 = optimize_stage_2_ess_hsh_scheduling(t, mg_data, s1, bess_soc, hydro_soc)
        logger.info(
            "Stage 2  status=%-10s  obj=%-10s  balance=%+.3f kW  "
            "BESS_SOC=%.1f kWh  H2_SOC=%.1f kWh",
            s2["solve_status"],
            f"{s2['objective_value']:.4f}" if s2["objective_value"] is not None else "N/A",
            s2["energy_balance"],
            s2["bess_soc"],
            s2["hyd_soc"],
        )

        # Update state
        bess_soc  = s2["bess_soc"]
        hydro_soc = s2["hyd_soc"]

        # Update prediction reliability factors based on actual vs predicted
        for i, df in mg_data.items():
            row = df.iloc[t]
            pv_err = abs(row["PV_Actual"] - row["PV_Predicted"]) / max(1e-3, row["PV_Predicted"])
            wt_err = abs(row["WT_Actual"] - row["WT_Predicted"]) / max(1e-3, row["WT_Predicted"])
            new_rf = 1 - 0.5 * (pv_err + wt_err)
            pred_reliability[i] = max(0.5, min(1.0, new_rf))

            # Update forward-looking factors (rolling look-ahead)
            next_t = min(t + 1, n_steps - 1)
            next_row = df.iloc[next_t]
            curr_gen = row["PV_Predicted"]    + row["WT_Predicted"]
            next_gen = next_row["PV_Predicted"] + next_row["WT_Predicted"]
            ff_new = next_gen / max(1e-3, curr_gen)
            forward_factor[i] = max(0.5, min(2.0, ff_new))

        # Combine and store
        combined = {
            "time":   timestamp,
            "step":   t,
            "stage1": s1,
            "stage2": s2,
        }
        history.append(combined)

        # ---- Plot ----
        fig, axes = create_real_time_dashboard(t, mg_data, s1, s2, history, fig, axes)
        if interactive_plot:
            plt.pause(TIME_STEP_DELAY)
        else:
            time.sleep(0.01)   # minimal delay when running non-interactively

        elapsed = time.time() - t_start
        logger.info("Step %d completed in %.2f s", t, elapsed)

    # ---- Save results ----
    save_results_to_excel(history, EXCEL_OUTPUT)

    if interactive_plot:
        plt.ioff()

    logger.info("=" * 70)
    logger.info("Optimization complete.  Results: %s", OUTPUT_DIR)
    logger.info("=" * 70)
    return history


# ============================================================
# SECTION 9 – ENTRY POINT
# ============================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Integrated Two-Stage Microgrid Optimization"
    )
    parser.add_argument(
        "--steps", type=int, default=None,
        help="Maximum number of time steps to run (default: all)",
    )
    parser.add_argument(
        "--interactive", action="store_true",
        help="Show live matplotlib window (requires a display)",
    )
    args = parser.parse_args()

    run_optimization(max_steps=args.steps, interactive_plot=args.interactive)
